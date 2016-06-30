#-*- coding: utf-8 -*-
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Font
import os
import datetime
class PipelineWorkbook():
	pass


class PipelineWorkshop():
	def get_file_names(self):
		cur_dir = os.path.dirname(os.path.abspath(__file__))
		input_dir = cur_dir
		for root, dirs, files in os.walk(input_dir, topdown=False):
		    files = [os.path.join(root,name) for name in files if name.find('.xlsx') != -1]
		return files

	def find_row_excld_title(self,ws):
		#找到第一个内容为"序号　No."的行号
		for i in range(1,ws.max_row+1):
			if ws.cell(row = i,column = 1).value == u"序号 No.":
				return i
		return 1

	def go(self):
		files = self.get_file_names()
		sum_wb = Workbook()
		sum_ws = sum_wb.active
		for file in files:
			p = PipelineWorkbook()
			p.wb_dir = os.path.dirname(file)
			p.wb_name = os.path.basename(file)
			wb = load_workbook(file)
			ws = wb.active
			p.ws_name = ws.title
			p.ws_range = ws.calculate_dimension()
			p.ws_first_row_excld_title = self.find_row_excld_title(ws) + 2
			p.max_row = ws.max_row


			from_start_row = p.ws_first_row_excld_title
			from_start_col = 1
			from_end_row = p.max_row
			from_end_col = 18
			to_start_row = sum_ws.max_row + 1
			to_start_col = 3

			print p.wb_dir,p.wb_name,p.ws_name,p.ws_range

			self.copy_range(
				from_ws=ws,
				to_ws=sum_ws,
				from_start_row = p.ws_first_row_excld_title,
				from_start_col = 1,
				from_end_row = p.max_row,
				from_end_col = 18,
				to_start_row = sum_ws.max_row + 1,
				to_start_col = 3)
			
			#设定第一列为excel文件名
			#设定第二列为worksheet名字
			for i in range(to_start_row,to_start_row + from_end_row - from_start_row + 1):
				sum_ws.cell(row = i,column = 1).value = p.wb_name 
				sum_ws.cell(row = i,column = 2).value = p.ws_name
		#在总表第一行加入标题行
		titles =u"销售员,工作表名,SR,date,项目跟踪,project,现状,客户,产品分类 Products Line,电压等级 Voltage,产品描述 Description,数量,预计金额,招标时间,预计发货,设计院 designing institute,设计师 designer,行动计划和拜访记录 action plan&visit plan,中标厂家 winner,丢标原因 reason".split(",")
		self.add_row(sum_ws,1,titles)
		self.format_work_sheet(sum_ws)
		self.check_error(sum_ws)

		self.save_wb(sum_wb)
	
	def get_timed_name(self,common_part=u"汇总",last = ".xlsx"):
		return common_part + datetime.datetime.now().strftime('%y%m%d_%H-%M-%S') +last
	def save_wb(self,wb,sub_folder = u'汇总结果'):
		target_path = os.path.join(os.path.dirname(__file__),sub_folder)
		file_name = self.get_timed_name()
		full_name = os.path.join(target_path,file_name)
		if not os.path.exists(target_path):
			os.mkdir(target_path)
		wb.save(full_name)
		print "Saved to ",full_name


	def copy_range(self,from_ws,to_ws,from_start_row,from_start_col,from_end_row,from_end_col,to_start_row,to_start_col):
		row_num = from_end_row - from_start_row + 1
		col_num = from_end_col - from_start_col + 11
		for cnt_row in range(row_num):
			for cnt_col in range(col_num):
				to_ws.cell(row = to_start_row + cnt_row,column = to_start_col + cnt_col).value = \
				from_ws.cell(row = from_start_row + cnt_row,column = from_start_col + cnt_col).value
				# self.format_cell(to_ws.cell(row = to_start_row + cnt_row,column = to_start_col + cnt_col))
		print 'copy from (row,col) = (%s,%s):(%s,%s) to (%s,%s)' %(from_start_row,from_start_col,from_end_row,from_end_col,to_start_row,to_start_col)
	
	def format_cell(self,cell,type = 'normal'):
		# my_font = Font(size = 8)
		my_font = Font(name='Calibri',
			size=8,
			bold=False,
			italic=False,
			vertAlign=None,
			underline='none',
			strike=False,
			color='FF000000'
			# color='FFBB00'
			)
		if type == "error":
			my_font.color = 'F60'
			my_font.size = 11

		cell.font = my_font
	def format_work_sheet(self,ws):
		for row in ws.iter_rows(range_string = ws.calculate_dimension()):
			for cell in row:
				self.format_cell(cell)

	def add_row(self,ws,row_num,data_list,offset = 0):
		i = 0
		for data in data_list:
			ws.cell(row = row_num,column = 1+offset+i).value = data
			i += 1

	def find_sr_form_filename(self,filename):
		srs = u"毕海滨,陈波,陈宇峰,傅强,顾问宇,韩焕贤,李强,徐政,杨俊,张宏军,周川一,周建国,周毅佳".split(',')
		for sr in srs:
			if filename.find(sr) != -1:
				return sr
		return filename

	def check_error(self,ws,first_data_row = 2):
		max_col = ws.max_column
		max_row = ws.max_row
		for col_cnt in range(1,ws.max_column + 1):
			for row_cnt in range(first_data_row,ws.max_row +1):
				cell = ws.cell(row = row_cnt,column = col_cnt)
				if col_cnt == 1:#文件名列
					cell.value = self.find_sr_form_filename(cell.value)
				elif col_cnt == 13:#金额列
					if cell.value:
						if not isinstance(cell.value,(long,int,float)):
							#金额列不为数字错误
							cell.font = Font(color = "FFFF0000",size = 11)
						elif not (cell.value > 0 and cell.value <= 10000):
							#金额单位可能不为万元错误
							cell.font = Font(color = "FF00FF00",size = 11)
					else:
						#金额列为空白错误
						cell.value = u"空白,建议删除"
						cell.font = Font(color = "FF0000ff",size = 11)
				elif col_cnt == 9:#产品分类
					if cell.value not in u"高压附件 HVA,高压电缆 HV,中压电缆 MV,中压附件 MVA".split(','):
						cell.font = Font(color = "FFFF0000",size = 11)
	


	# def test_copy(self,ws):
	# 	for i in range(1,10):
	# 		for j in range(1,10):
	# 			ws.cell(row=i,column=j).value = i*j

	# def test_save(self):
	# 	wb = Workbook()
	# 	ws = wb.active
	# 	# self.test_copy(ws)
	# 	# wb.save()
	# 	self.copy_range(from_ws=ws,to_ws=ws,from_start_row=1,from_start_col=1,from_end_row=9,from_end_col=9,to_start_row=1,to_start_col=1)
	# 	sum_file_name = os.path.join(os.path.dirname(__file__),self.get_timed_name('test_save'))
	# 	wb.save(sum_file_name)
	# 	print 'saved to ',sum_file_name
	# def test(self):
	# 	# wb = load_workbook()
	# 	wb = Workbook()
	# 	ws = wb.active
	# 	for i in range(1,10):
	# 		for j in range(1,10):
	# 			ws.cell(row = i ,column = j).value = 'row%s col%s' %(i,j)

	# 	for col in ws.iter_cols(min_row = 1,max_row = ws.max_row,min_col = 1,max_col = ws.max_column):
	# 		for cell in col:
	# 			print cell.value


p = PipelineWorkshop()
p.go()
print u"完成,请关闭"
raw_input()
# p.test()



